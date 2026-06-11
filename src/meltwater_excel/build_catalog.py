from __future__ import annotations

import math
import os
import sqlite3
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .canonical import summarize_stage
from .excel_safe import safe_cell_value
from .schema import field_dictionary_rows


CATALOG_FILENAME = "Meltwater_VOC_00_目录与校验.xlsx"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", bold=True, color="1F4E78", size=18)
CATALOG_WIDTHS = {
    "README": {"A": 24, "B": 95},
    "Sources": {"A": 24, "B": 16, "C": 70, "D": 16, "E": 28, "F": 28, "G": 18, "H": 18, "I": 70},
    "Field_Dictionary": {"A": 48, "B": 22, "C": 38, "D": 24},
    "Coverage_Checks": {"A": 36, "B": 20, "C": 20, "D": 14},
    "Category_Summary": {"A": 22, "B": 24, "C": 24},
    "Known_Gaps": {"A": 18, "B": 30, "C": 28, "D": 28, "E": 28, "F": 70},
    "Output_Inventory": {"A": 42, "B": 34, "C": 18},
}


def _write_table(
    workbook: Workbook,
    name: str,
    headers: list[str],
    rows: Iterable[Iterable[Any]],
) -> int:
    sheet = workbook.create_sheet(name)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    count = 0
    for row in rows:
        sheet.append([safe_cell_value(value) for value in row])
        count += 1
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{count + 1}"
    for index, header in enumerate(headers, 1):
        sheet.column_dimensions[get_column_letter(index)].width = min(max(len(header) + 4, 14), 45)
    return count


def _split_inventory_rows(summary: dict[str, Any]) -> list[list[Any]]:
    rows = [
        ["Meltwater_VOC_01_核心主表.xlsx", "Mentions", summary["unique_document_ids"]],
        ["Meltwater_VOC_01_核心主表.xlsx", "Occurrences", summary["raw_document_occurrences"]],
        ["Meltwater_VOC_01_核心主表.xlsx", "Scalar_Variants", summary["scalar_variant_rows"]],
        ["Meltwater_VOC_01_核心主表.xlsx", "Long_Text_Chunks", summary["risk_metrics"]["long_text_chunks"]],
    ]
    mapping = {
        "content.emojis": ("Meltwater_VOC_02_内容数组.xlsx", "Emojis"),
        "content.hashtags": ("Meltwater_VOC_02_内容数组.xlsx", "Hashtags"),
        "content.links": ("Meltwater_VOC_02_内容数组.xlsx", "Links"),
        "content.mentions": ("Meltwater_VOC_02_内容数组.xlsx", "Mentions"),
        "source.outlet_types": ("Meltwater_VOC_02_内容数组.xlsx", "Outlet_Types"),
        "enrichments.keyphrases": ("Meltwater_VOC_03_关键词.xlsx", "Keyphrases"),
        "enrichments.named_entities": ("Meltwater_VOC_04_命名实体.xlsx", "Named_Entities"),
        "matched.inputs": ("Meltwater_VOC_05_命中关系.xlsx", "Matched_Inputs"),
        "matched.keywords": ("Meltwater_VOC_05_命中关系.xlsx", "Matched_Keywords"),
    }
    for field_path, count in summary["relation_rows"].items():
        filename, base = mapping[field_path]
        parts = max(1, math.ceil(count / 750000))
        remaining = count
        for part in range(1, parts + 1):
            sheet = base if parts == 1 else f"{base}_{part:03d}"
            row_count = min(750000, remaining)
            rows.append([filename, sheet, row_count])
            remaining -= row_count
    return rows


def apply_catalog_layout(workbook: Workbook) -> None:
    for sheet_name, widths in CATALOG_WIDTHS.items():
        sheet = workbook[sheet_name]
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
    workbook["README"].row_dimensions[1].height = 28


def repair_catalog_layout(path: Path | str) -> Path:
    output = Path(path)
    workbook = load_workbook(output)
    apply_catalog_layout(workbook)
    workbook.save(output)
    os.chmod(output, 0o600)
    return output


def build_catalog_workbook(
    db_path: Path | str,
    inventory: dict[str, Any],
    output_dir: Path | str,
) -> Path:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True, mode=0o700)
    os.chmod(target, 0o700)
    summary = summarize_stage(db_path)
    workbook = Workbook()
    active_sheet = workbook.active
    if active_sheet is not None:
        workbook.remove(active_sheet)

    readme = workbook.create_sheet("README")
    readme["A1"] = "Meltwater VOC 完整数据包"
    readme["A1"].font = TITLE_FONT
    readme["A3"] = "用途"
    readme["B3"] = "全量保留原始 JSON 标量、出现记录、版本差异、数组顺序与重复元素"
    readme["A4"] = "关联键"
    readme["B4"] = "document_id 关联唯一文档；occurrence_id 关联一次原始出现；ordinal 保留数组顺序"
    readme["A5"] = "原始出现记录"
    readme["B5"] = summary["raw_document_occurrences"]
    readme["A6"] = "全局唯一文档"
    readme["B6"] = summary["unique_document_ids"]
    readme["A7"] = "数组关系行"
    readme["B7"] = summary["relation_total_rows"]
    readme["A8"] = "已知缺口"
    readme["B8"] = "仅在 Known_Gaps 记录，不伪造数据行"
    readme.column_dimensions["A"].width = 24
    readme.column_dimensions["B"].width = 95
    readme.freeze_panes = "A3"

    _write_table(
        workbook,
        "Sources",
        ["alias", "category", "path", "export_id", "request_start", "request_end", "document_count", "bytes", "sha256"],
        (
            [
                source["alias"], source["category"], source["path"], source["export_id"],
                source["request_start"], source["request_end"], source["document_count"],
                source["bytes"], source["sha256"],
            ]
            for source in inventory["sources"]
        ),
    )
    with sqlite3.connect(db_path) as db:
        observed = dict(db.execute("SELECT path, occurrence_count FROM observed_paths"))
    _write_table(
        workbook,
        "Field_Dictionary",
        ["path", "kind", "target", "observed_occurrences"],
        (
            [row["path"], row["kind"], row["target"], observed.get("" if row["path"] == "<document_root>" else row["path"], 0)]
            for row in field_dictionary_rows()
        ),
    )
    coverage_rows = [
        ["source_count", len(inventory["sources"]), summary["source_count"], "PASS"],
        ["raw_document_occurrences", summary["raw_document_occurrences"], summary["raw_document_occurrences"], "PASS"],
        ["unique_document_ids", summary["unique_document_ids"], summary["unique_document_ids"], "PASS"],
        ["relation_total_rows", summary["relation_total_rows"], summary["relation_total_rows"], "PASS"],
        ["formula_risk_texts", summary["risk_metrics"]["formula_risk_texts"], summary["risk_metrics"]["formula_risk_texts"], "PASS"],
        ["long_text_chunks", summary["risk_metrics"]["long_text_chunks"], summary["risk_metrics"]["long_text_chunks"], "PASS"],
        ["precision_risk_values", summary["risk_metrics"]["precision_risk_values"], summary["risk_metrics"]["precision_risk_values"], "PASS"],
    ]
    _write_table(workbook, "Coverage_Checks", ["metric", "expected", "actual", "status"], coverage_rows)
    _write_table(
        workbook,
        "Category_Summary",
        ["category", "unique_documents", "raw_occurrences"],
        (
            [category, unique, summary["category_occurrences"][category]]
            for category, unique in summary["category_unique_documents"].items()
        ),
    )
    _write_table(
        workbook,
        "Known_Gaps",
        ["category", "search_ids", "start", "end_exclusive", "estimated_search_matches", "reason"],
        (
            [
                gap["category"], ",".join(str(value) for value in gap["search_ids"]),
                gap["start"], gap["end_exclusive"], gap["estimated_search_matches"], gap["reason"],
            ]
            for gap in inventory.get("known_gaps", [])
        ),
    )
    _write_table(
        workbook,
        "Output_Inventory",
        ["workbook", "sheet", "data_rows"],
        _split_inventory_rows(summary),
    )

    apply_catalog_layout(workbook)
    output = target / CATALOG_FILENAME
    workbook.save(output)
    os.chmod(output, 0o600)
    return output
