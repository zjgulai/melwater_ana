from __future__ import annotations

import json
import os
import random
from pathlib import Path
from typing import Any, cast

from openpyxl import load_workbook

from .build_relations import OUTPUTS
from .excel_safe import safe_cell_value
from .inventory import iter_documents, load_source_config, resolve_source_path
from .schema import ARRAY_PATHS, get_path
from .staging import json_text


MISSING = object()


def _array_state(value: Any) -> dict[str, Any]:
    if value is MISSING:
        return {"state": "missing", "count": 0}
    if value is None:
        return {"state": "null", "count": 0}
    return {"state": "empty" if not value else "nonempty", "count": len(value)}


def _collect_expected(
    config: dict[str, Any],
    samples_per_source: int,
    seed: int,
) -> dict[str, dict[str, Any]]:
    root = Path(config["_project_root"])
    expected: dict[str, dict[str, Any]] = {}
    for source in config["sources"]:
        sample_size = min(samples_per_source, source["expected_documents"])
        # Deterministic audit sampling only, not security or cryptography.
        rng = random.Random(f"{seed}:{source['alias']}")  # nosec B311
        indices = set(rng.sample(range(source["expected_documents"]), sample_size))
        path = resolve_source_path(source["path"], root)
        for index, document in enumerate(iter_documents(path)):
            if index not in indices:
                continue
            occurrence_id = f"{source['alias']}:{index}"
            arrays = {}
            states = {}
            for field_path in ARRAY_PATHS:
                value = get_path(document, field_path, MISSING)
                states[field_path] = _array_state(value)
                arrays[field_path] = [] if value in (MISSING, None) else [
                    (ordinal, json_text(item)) for ordinal, item in enumerate(value)
                ]
            expected[occurrence_id] = {
                "document_id": str(document["id"]),
                "source_alias": source["alias"],
                "category": source["category"],
                "states": states,
                "arrays": arrays,
            }
    return expected


def _audit_occurrences(
    output_dir: Path,
    expected: dict[str, dict[str, Any]],
    failures: list[str],
) -> None:
    path = output_dir / "Meltwater_VOC_01_核心主表.xlsx"
    workbook = load_workbook(path, read_only=True, data_only=False)
    sheet = workbook["Occurrences"]
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    indexes = {header: index for index, header in enumerate(headers)}
    found = set()
    for row in rows:
        occurrence_id = str(row[indexes["occurrence_id"]])
        if occurrence_id not in expected:
            continue
        found.add(occurrence_id)
        item = expected[occurrence_id]
        for column in ["document_id", "source_alias", "category"]:
            if row[indexes[column]] != safe_cell_value(item[column]):
                failures.append(f"{occurrence_id}:{column}")
        for field_path in ARRAY_PATHS:
            state = item["states"][field_path]
            if row[indexes[f"{field_path}__state"]] != state["state"]:
                failures.append(f"{occurrence_id}:{field_path}:state")
            if row[indexes[f"{field_path}__count"]] != state["count"]:
                failures.append(f"{occurrence_id}:{field_path}:count")
    workbook.close()
    for occurrence_id in expected.keys() - found:
        failures.append(f"{occurrence_id}:missing_occurrence")


def _audit_relations(
    output_dir: Path,
    expected: dict[str, dict[str, Any]],
    failures: list[str],
) -> int:
    actual: dict[str, dict[str, list[tuple[int, Any]]]] = {
        occurrence_id: {field_path: [] for field_path in ARRAY_PATHS}
        for occurrence_id in expected
    }
    relation_item_count = 0
    for filename, relations in OUTPUTS.items():
        workbook = load_workbook(output_dir / filename, read_only=True, data_only=False)
        for sheet in workbook.worksheets:
            field_path = next(
                field for field, base in relations
                if sheet.title == base or sheet.title.startswith(f"{base}_")
            )
            rows = sheet.iter_rows(values_only=True)
            headers = list(next(rows))
            indexes = {header: index for index, header in enumerate(headers)}
            for row in rows:
                occurrence_id = str(row[indexes["occurrence_id"]])
                if occurrence_id not in expected:
                    continue
                ordinal = cast(int | float | str, row[indexes["ordinal"]])
                actual[occurrence_id][field_path].append(
                    (int(ordinal), row[indexes["item_json"]])
                )
                relation_item_count += 1
        workbook.close()
    for occurrence_id, item in expected.items():
        for field_path in ARRAY_PATHS:
            if actual[occurrence_id][field_path] != item["arrays"][field_path]:
                failures.append(f"{occurrence_id}:{field_path}:items")
    return relation_item_count


def audit_random_samples(
    config_path: Path | str,
    output_dir: Path | str,
    samples_per_source: int = 20,
    seed: int = 20260604,
    output: Path | str | None = None,
) -> dict[str, Any]:
    config = load_source_config(config_path)
    target = Path(output_dir)
    expected = _collect_expected(config, samples_per_source, seed)
    failures: list[str] = []
    _audit_occurrences(target, expected, failures)
    relation_item_count = _audit_relations(target, expected, failures)
    result = {
        "status": "PASS" if not failures else "FAIL",
        "seed": seed,
        "samples_per_source": samples_per_source,
        "sample_occurrences": len(expected),
        "sample_relation_items": relation_item_count,
        "failed_checks": failures,
    }
    if output:
        path = Path(output)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        os.chmod(path, 0o600)
    if failures:
        raise ValueError(f"sample audit failed: {', '.join(failures[:10])}")
    return result
