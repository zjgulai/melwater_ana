from __future__ import annotations

import math
import os
import zipfile
from pathlib import Path
from typing import Any, Iterable

from defusedxml import ElementTree as ET
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .excel_safe import safe_cell_value
from .inventory import sha256_file


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


class StreamingWorkbook:
    def __init__(self, output: Path | str, max_data_rows: int = 750000):
        self.output = Path(output)
        self.max_data_rows = max_data_rows
        self.workbook = Workbook(write_only=True)
        self.sheet_rows: dict[str, int] = {}

    def _new_sheet(self, name: str, headers: list[str]):
        sheet = self.workbook.create_sheet(name[:31])
        sheet.freeze_panes = "A2"
        for index, header in enumerate(headers, 1):
            width = min(max(len(header) + 3, 12), 35)
            sheet.column_dimensions[get_column_letter(index)].width = width
        header_cells = []
        for header in headers:
            cell = WriteOnlyCell(sheet, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            header_cells.append(cell)
        sheet.append(header_cells)
        return sheet

    def write_family(
        self,
        base_name: str,
        headers: list[str],
        rows: Iterable[Iterable[Any]],
        total_rows: int,
    ) -> dict[str, int]:
        parts = max(1, math.ceil(total_rows / self.max_data_rows))
        iterator = iter(rows)
        written_total = 0
        result: dict[str, int] = {}
        for part in range(1, parts + 1):
            name = base_name if parts == 1 else f"{base_name}_{part:03d}"
            sheet = self._new_sheet(name, headers)
            part_count = min(self.max_data_rows, total_rows - written_total)
            for _ in range(part_count):
                try:
                    row = next(iterator)
                except StopIteration as exc:
                    raise ValueError(
                        f"{base_name} expected {total_rows} rows, stopped at {written_total}"
                    ) from exc
                sheet.append([safe_cell_value(value) for value in row])
                written_total += 1
            sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{part_count + 1}"
            result[name[:31]] = part_count + 1
        try:
            next(iterator)
        except StopIteration:
            pass
        else:
            raise ValueError(f"{base_name} produced more than {total_rows} rows")
        self.sheet_rows.update(result)
        return result

    def save(self) -> Path:
        self.output.parent.mkdir(parents=True, exist_ok=True, mode=0o700)
        os.chmod(self.output.parent, 0o700)
        self.workbook.save(self.output)
        os.chmod(self.output, 0o600)
        return self.output


def _relationship_targets(archive: zipfile.ZipFile) -> dict[str, str]:
    root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    return {
        relation.attrib["Id"]: relation.attrib["Target"]
        for relation in root
    }


def inspect_workbook(path: Path | str) -> dict[str, Any]:
    workbook_path = Path(path)
    sheets: dict[str, int] = {}
    formula_count = 0
    max_text_length = 0
    with zipfile.ZipFile(workbook_path) as archive:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        targets = _relationship_targets(archive)
        namespace = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        relation_key = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
        sheets_element = workbook_root.find("m:sheets", namespace)
        if sheets_element is None:
            raise ValueError(f"{workbook_path} missing workbook sheets metadata")
        for sheet in sheets_element:
            name = sheet.attrib["name"]
            target = targets[sheet.attrib[relation_key]]
            normalized = target.lstrip("/")
            if not normalized.startswith("xl/"):
                normalized = f"xl/{normalized}"
            rows = 0
            with archive.open(normalized) as handle:
                for _, element in ET.iterparse(handle, events=("end",)):
                    tag = element.tag.rsplit("}", 1)[-1]
                    if tag == "row":
                        rows += 1
                    elif tag == "f":
                        formula_count += 1
                    elif tag == "t" and element.text:
                        max_text_length = max(max_text_length, len(element.text))
                    element.clear()
            sheets[name] = rows
        if "xl/sharedStrings.xml" in archive.namelist():
            with archive.open("xl/sharedStrings.xml") as handle:
                for _, element in ET.iterparse(handle, events=("end",)):
                    if element.tag.rsplit("}", 1)[-1] == "t" and element.text:
                        max_text_length = max(max_text_length, len(element.text))
                    element.clear()
    return {
        "file": workbook_path.name,
        "bytes": workbook_path.stat().st_size,
        "sha256": sha256_file(workbook_path),
        "sheets": sheets,
        "formula_count": formula_count,
        "max_cell_text_length": max_text_length,
    }
