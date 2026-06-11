#!/usr/bin/env python3
"""Meltwater VOC 数据采集脚本。

用法:
    python3 scripts/collect.py 消毒器
    python3 scripts/collect.py 暖奶器 --start 2026-01-01 --end 2026-03-31
    python3 scripts/collect.py 吸奶器 --days 90
    python3 scripts/collect.py --list
"""
import argparse
import json
import os
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import urlparse
from urllib.request import Request, urlopen
from urllib.error import HTTPError

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

CONFIG_PATH = PROJECT_ROOT / "config" / "categories.json"
DATA_EXPORTS = PROJECT_ROOT / "data" / "exports"
DATA_EXCEL = PROJECT_ROOT / "data" / "excel"
MAX_SEARCHES_PER_EXPORT = 5
LIVE_API_ENV_VAR = "LEGACY_COLLECT_LIVE"
LIVE_API_EXIT_CODE = 2

DATA_EXPORTS.mkdir(parents=True, exist_ok=True)
DATA_EXCEL.mkdir(parents=True, exist_ok=True)


def legacy_live_enabled(environ=None):
    env = os.environ if environ is None else environ
    return env.get(LIVE_API_ENV_VAR) == "1"


def require_legacy_live_enabled(environ=None):
    if legacy_live_enabled(environ):
        return
    print(
        "❌ scripts/collect.py 是旧版 live API 采集脚本，默认禁止联网执行。\n"
        f"   如确需使用旧链路，请在本次命令前显式设置 {LIVE_API_ENV_VAR}=1。\n"
        "   推荐优先使用 manifest 驱动的新数据链路，并记录本次采集范围。",
        file=sys.stderr,
    )
    sys.exit(LIVE_API_EXIT_CODE)


def load_env():
    env_path = PROJECT_ROOT / ".env"
    if not env_path.exists():
        print("❌ 缺少 .env 文件，请创建并设置 MELTWATER_API_KEY")
        sys.exit(1)
    for line in env_path.read_text().strip().split("\n"):
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())


def load_config():
    with open(CONFIG_PATH) as f:
        return json.load(f)


def require_https_url(url):
    parsed = urlparse(url)
    if parsed.scheme != "https":
        raise ValueError(f"only HTTPS URLs are allowed: {url}")
    return url


def api_request(method, path, body=None):
    api_key = os.environ.get("MELTWATER_API_KEY", "")
    base = os.environ.get("MELTWATER_BASE_URL", "https://api.meltwater.com")
    url = require_https_url(f"{base}{path}")
    headers = {"Accept": "application/json", "apikey": api_key}
    data = json.dumps(body).encode() if body else None
    if body:
        headers["Content-Type"] = "application/json"
    req = Request(url, data=data, headers=headers, method=method)
    try:
        # URL scheme is validated as HTTPS before the request is created.
        with urlopen(req) as resp:  # nosec B310
            return json.loads(resp.read()), resp.status
    except HTTPError as e:
        body = e.read().decode()
        print(f"  API错误 {e.code}: {body[:300]}")
        return None, e.code


def list_categories(config):
    print("可用品类：")
    for name, cfg in config["categories"].items():
        ids = cfg["search_ids"]
        print(f"  {name:<10} → {len(ids)}个搜索 {ids}")


def create_exports(category_name, search_ids, start, end):
    batched = [
        search_ids[i : i + MAX_SEARCHES_PER_EXPORT]
        for i in range(0, len(search_ids), MAX_SEARCHES_PER_EXPORT)
    ]
    exports = []
    for batch in batched:
        body = {
            "onetime_export": {
                "search_ids": batch,
                "start_date": f"{start}T00:00:00Z",
                "end_date": f"{end}T00:00:00Z",
                "template": {"name": "api.json"},
            }
        }
        result, code = api_request("POST", "/v3/exports/one-time", body)
        if result and "onetime_export" in result:
            exp = result["onetime_export"]
            exports.append(exp)
            searches = [s["name"] for s in exp["searches"]]
            print(f"  导出 {exp['id']} 已创建 → {', '.join(searches)}")
        else:
            print(f"  导出创建失败: {result}")
    return exports


def wait_exports(exports):
    pending = [e for e in exports]
    finished = []
    max_wait = 600  # 10分钟
    waited = 0
    while pending and waited < max_wait:
        time.sleep(15)
        waited += 15
        still_pending = []
        for exp in pending:
            result, _ = api_request("GET", f"/v3/exports/one-time/{exp['id']}")
            if result and result.get("onetime_export", {}).get("status") == "FINISHED":
                print(f"  导出 {exp['id']} 完成 ✅")
                finished.append(result["onetime_export"])
            else:
                still_pending.append(exp)
        pending = still_pending
        if pending:
            print(f"  等待中... ({len(pending)}个, 已等{waited}秒)")
    if pending:
        print(f"  ⚠️ {len(pending)}个导出超时未完成")
    return finished


def download_export(exp, category_name):
    data_url = require_https_url(exp["data_url"])
    exp_id = exp["id"]
    req = Request(data_url, headers={"apikey": os.environ["MELTWATER_API_KEY"]})
    # URL scheme is validated as HTTPS before the request is created.
    with urlopen(req) as resp:  # nosec B310
        data = json.loads(resp.read())
    filename = f"{category_name}_{exp_id}_{datetime.now().strftime('%Y%m%d')}.json"
    filepath = DATA_EXPORTS / filename
    with open(filepath, "w") as f:
        json.dump(data, f, ensure_ascii=False)
    return filepath, data


def json_to_excel_rows(data, category_name):
    docs = data.get("documents", data.get("docs", []))
    rows = []
    for doc in docs:
        rows.append(flatten_doc(category_name, doc))
    return rows


def flatten_doc(category, doc):
    src = doc.get("source") or {}
    mtr = src.get("metrics") or {}
    auth = doc.get("author") or {}
    cont = doc.get("content") or {}
    enr = doc.get("enrichments") or {}
    loc = doc.get("location") or {}
    eng = (doc.get("metrics") or {}).get("engagement") or {}
    se = (doc.get("metrics") or {}).get("social_echo") or {}
    cus = doc.get("custom") or {}
    mat = doc.get("matched") or {}
    thr = doc.get("thread") or {}
    par = doc.get("parent") or {}

    def s(v):
        if v is None:
            return ""
        if isinstance(v, list):
            return "; ".join(str(x) for x in v if x)
        return str(v).strip()

    def n(v):
        if v is None:
            return ""
        try:
            return float(v)
        except (TypeError, ValueError):
            return str(v)

    ne = "; ".join(f"{e.get('name','')}({e.get('sentiment','')})" for e in (enr.get("named_entities") or []) if e.get("name"))
    ms = "; ".join(m.get("name","") for m in (mat.get("inputs") or []) if m.get("name"))
    cc = "; ".join(c.get("name","") for c in (cus.get("custom_categories") or []) if c.get("name"))

    return [
        category, s(doc.get("id")), s(doc.get("published_date")), s(doc.get("indexed_date")),
        s(doc.get("url")), s(doc.get("content_type")), s(src.get("information_type")),
        s(src.get("type")), s(src.get("name")), s(src.get("domain")), s(src.get("url")),
        s(auth.get("name")), s(auth.get("handle")), s(auth.get("profile_url")),
        s(cont.get("title")), s(cont.get("opening_text")), s(cont.get("body")),
        s(cont.get("hashtags")), s(cont.get("mentions")), s(cont.get("emojis")),
        s(enr.get("sentiment")), s(enr.get("language_code")), s(loc.get("country_code")),
        s(loc.get("city")), s(loc.get("state")), ne, "; ".join(enr.get("keyphrases") or []),
        n(mtr.get("reach")), n(mtr.get("ave")), n(eng.get("total")), n(eng.get("likes")),
        n(eng.get("replies")), n(eng.get("comments")), n(eng.get("shares")),
        n(eng.get("reactions")), n((doc.get("metrics") or {}).get("views")),
        n(se.get("total")), n((doc.get("metrics") or {}).get("editorial_echo")),
        n((doc.get("metrics") or {}).get("earned_media_value")), ms,
        "; ".join(mat.get("keywords") or []), s(mat.get("hit_sentence")),
        "; ".join(cus.get("tags") or []), cc, s(thr.get("title")), s(thr.get("url")), s(par.get("url")),
    ]


HEADERS = [
    "品类", "文档ID", "发布时间", "入库时间", "URL", "内容类型", "信息类型",
    "来源类型", "来源名称", "来源域名", "来源URL", "作者名", "作者Handle", "作者主页",
    "标题", "正文开头", "正文", "Hashtags", "Mentions", "Emojis",
    "情感", "语言", "国家码", "城市", "州/省", "命名实体", "关键词",
    "来源触达量", "AVE", "总互动量", "点赞数", "回复数", "评论数", "分享数", "反应数",
    "浏览量", "社交回声", "编辑回声", "EMV", "命中搜索", "命中关键词", "命中句子",
    "标签", "自定义分类", "讨论串标题", "讨论串URL", "父级URL",
]


def create_excel(rows, category_name, date_str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"{category_name}_VOC"

    hf = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bd = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
    cf = Font(name="微软雅黑", size=10)

    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, halign, bd

    for ri, row in enumerate(rows, 2):
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=v if v != "" else None)
            c.font, c.border = cf, bd
            if ci == 1:
                c.alignment = Alignment(horizontal="center", vertical="center")

    widths = {1:14,2:36,3:20,4:20,5:50,6:16,7:10,8:14,9:40,10:30,11:50,12:20,13:22,
              14:50,15:60,16:60,17:60,18:25,19:25,20:20,21:10,22:20,23:8,24:15,25:15,
              26:40,27:40,28:12,29:10,30:10,31:10,32:10,33:10,34:10,35:10,36:10,37:10,
              38:10,39:30,40:30,41:50,42:25,43:25,44:40,45:40}
    for ci, w in widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    safe = category_name.replace("(", "").replace(")", "").replace(" ", "_")
    path_full = DATA_EXCEL / f"Meltwater_{safe}_{date_str}.xlsx"
    wb.save(path_full)
    print(f"  → {path_full} ({len(rows):,}条)")

    neg_rows = [r for r in rows if r[20] == "negative"]
    if neg_rows:
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = f"{category_name}_负面VOC"
        for ci, h in enumerate(HEADERS, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.font, c.fill, c.alignment, c.border = hf, hfill, halign, bd
        for ri, row in enumerate(neg_rows, 2):
            for ci, v in enumerate(row, 1):
                c = ws2.cell(row=ri, column=ci, value=v if v != "" else None)
                c.font, c.border = cf, bd
        for ci, w in widths.items():
            ws2.column_dimensions[get_column_letter(ci)].width = w
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
        path_neg = DATA_EXCEL / f"Meltwater_{safe}_负面VOC_{date_str}.xlsx"
        wb2.save(path_neg)
        print(f"  → {path_neg} ({len(neg_rows):,}条 负面)")


def print_summary(rows, category_name, start, end):
    from collections import Counter
    sentiments = Counter(r[20] for r in rows)
    total = len(rows)
    pos = sentiments.get("positive", 0)
    neg = sentiments.get("negative", 0)
    neu = sentiments.get("neutral", 0)
    days = max((datetime.fromisoformat(end) - datetime.fromisoformat(start)).days, 1)
    print(f"\n{'='*50}")
    print(f"  {category_name} 采集完成")
    print(f"  时间: {start} ~ {end} ({days}天)")
    print(f"  总量: {total:,}条 | 日均: {total//days}")
    print(f"  情感: +{pos}({pos*100/max(total,1):.1f}%) / -{neg}({neg*100/max(total,1):.1f}%) / N{neu}({neu*100/max(total,1):.1f}%)")
    print(f"{'='*50}")


def main():
    config = load_config()

    parser = argparse.ArgumentParser(description="Meltwater VOC 数据采集")
    parser.add_argument("category", nargs="?", help="品类名称（如 消毒器/暖奶器/吸奶器）")
    parser.add_argument("--start", help="开始日期 YYYY-MM-DD（默认30天前）")
    parser.add_argument("--end", help="结束日期 YYYY-MM-DD（默认今天）")
    parser.add_argument("--days", type=int, default=30, help="采集最近N天（默认30）")
    parser.add_argument("--list", action="store_true", help="列出可用品类")
    parser.add_argument("--skip-excel", action="store_true", help="仅下载JSON，不生成Excel")
    args = parser.parse_args()

    if args.list or not args.category:
        list_categories(config)
        return

    category_name = args.category
    if category_name not in config["categories"]:
        print(f"❌ 未知品类: {category_name}")
        print(f"可用: {', '.join(config['categories'].keys())}")
        sys.exit(1)

    require_legacy_live_enabled()
    load_env()

    cfg = config["categories"][category_name]
    search_ids = cfg["search_ids"]
    end_date = args.end or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    start_date = args.start or (datetime.now(timezone.utc) - timedelta(days=args.days)).strftime("%Y-%m-%d")

    print(f"\n🔍 采集品类: {category_name}")
    print(f"  搜索数: {len(search_ids)} | 时间: {start_date} ~ {end_date}")
    print(f"  描述: {cfg['description']}")

    exports = create_exports(category_name, search_ids, start_date, end_date)
    if not exports:
        print("❌ 导出创建失败")
        sys.exit(1)

    print("\n⏳ 等待导出完成...")
    finished = wait_exports(exports)

    all_rows = []
    for exp in finished:
        print(f"\n📥 下载导出 {exp['id']}...")
        filepath, data = download_export(exp, category_name)
        print(f"  → {filepath}")
        rows = json_to_excel_rows(data, category_name)
        all_rows.extend(rows)

    if not all_rows:
        print("❌ 无数据")
        sys.exit(1)

    print_summary(all_rows, category_name, start_date, end_date)

    if not args.skip_excel:
        print("\n📊 生成Excel...")
        date_str = f"{start_date}_{end_date}"
        create_excel(all_rows, category_name, date_str)

    print("\n✅ 完成!")


if __name__ == "__main__":
    main()
